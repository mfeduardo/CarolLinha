from PyQt6 import QtWidgets, uic
from PyQt6.QtWidgets import QMessageBox, QFileDialog
from openpyxl import load_workbook
import os, time, csv, clipboard

# váriáveis globais
arquivo = '-'
arquivo_saida = '-'
dados_config = []


def definir_configurcao():
    with open('arquivos/configuracao/config.csv') as configuracao:
        for registro in csv.reader(configuracao):
            global dados_config
            dados_config.append(registro[0])


definir_configurcao()


def alert(msg):
    info = QMessageBox()
    info.setIcon(QMessageBox.Icon.Warning)
    info.setWindowTitle('Atenção!')
    info.setText(f'{msg}         ')
    info.exec()


def open_file():
    global arquivo, arquivo_saida
    file_filter = 'Excel File (*.xlsx *.xls)'
    response = QFileDialog.getOpenFileName(
        filter=file_filter, caption='Selecione a planilha de origem', directory=arquivo)
    if response[0]:
        arquivo = response[0]
        main.lb_file_name.setText(arquivo)
        arquivo_saida, extensao = os.path.splitext(arquivo)
        arquivo_saida += ' - karina.xlsx'
        main.lb_save_name.setText(arquivo_saida)
        main.bt_edit_plan.setEnabled(True)
    else:
        main.lb_file_name.setText(arquivo)
        main.lb_save_name.setText(arquivo_saida)
        if not arquivo:
            main.bt_edit_plan.setEnabled(False)


def save_file():
    global arquivo_saida
    file_filter = 'Excel File (*.xlsx *.xls)'
    response = QFileDialog.getSaveFileName(
        filter=file_filter, caption='Selecione o arquivo', directory=arquivo_saida)
    if response[0]:
        arquivo_saida = response[0]
        arquivo_saida, extensao = os.path.splitext(arquivo_saida)
        arquivo_saida += '.xlsx'
        main.lb_save_name.setText(arquivo_saida)


def processar():
    if (arquivo != '-' and arquivo_saida):
        main.bt_process.setText('Aguarde...')
        gerar_recibo(arquivo, arquivo_saida)
    else:
        msg = 'Planilha de entrada não informada!'
        alert(msg)


def reset():
    global arquivo, arquivo_saida
    arquivo = '-'
    arquivo_saida = '-'
    main.lb_save_name.setText(arquivo_saida)
    main.lb_file_name.setText(arquivo_saida)
    main.bt_edit_plan.setEnabled(False)


def editar_configuracao():
    edit.le_aba_inicial.setText(dados_config[0])
    edit.le_aba_final.setText(dados_config[1])
    edit.le_col_inicial.setText(dados_config[2])
    edit.le_col_final.setText(dados_config[3])
    edit.le_linha_inicial.setText(dados_config[4])
    edit.le_linha_final.setText(dados_config[5])
    edit.le_linha_del.setText(dados_config[6])
    edit.show()


def salvar_configuracao():
    definicoes = [
        edit.le_aba_inicial.text(),
        edit.le_aba_final.text(),
        edit.le_col_inicial.text(),
        edit.le_col_final.text(),
        edit.le_linha_inicial.text(),
        edit.le_linha_final.text(),
        edit.le_linha_del.text()
    ]

    check = True
    for d in definicoes:
        if not d.isdigit():
            check = False

    if check:

        with open('arquivos/configuracao/config.csv', 'w') as configuracao:
            for registro in definicoes:
                print(f'{registro},', file=configuracao)

        global dados_config
        dados_config = definicoes
        definir_configurcao()

        edit.close()

    else:
        alert('- Preencha todos os campos!\n- Informe apenas números!')


def close():
    main.close()


# Remover espaços em branco nos parágrafos da planilha
def gerar_recibo(entrada, saida):
    arquivo_entrada = entrada
    arquivo_saida = saida

    file_path = arquivo_entrada
    workbook = load_workbook(filename=file_path)

    # Índice das abas
    start_sheet_index = int(dados_config[0])
    end_sheet_index = int(dados_config[1])
    if end_sheet_index > len(workbook.sheetnames):
        end_sheet_index = len(workbook.sheetnames)

    # remover espaços
    def remover_espacos(inicio, fim):
        for coluna in range(int(dados_config[2]), int(dados_config[3])+1):
            for row in sheet.iter_rows(min_row=inicio, max_row=fim, min_col=coluna, max_col=coluna):
                if type(row[0].value) == str:
                    row[0].value = ' '.join(row[0].value.split())
        return

    main.progressBar.show()
    main.progressBar.setMinimum(start_sheet_index)
    main.progressBar.setMaximum(end_sheet_index)

    for sheet_index in range(start_sheet_index, end_sheet_index):
        sheet = workbook.worksheets[sheet_index]

        time.sleep(0.01)
        main.progressBar.setValue(sheet_index+1)
        last_row_with_content = int(dados_config[6])

        # Remover linhas vazias abaixo da última linha com conteúdo
        if last_row_with_content < sheet.max_row:
            sheet.delete_rows(last_row_with_content + 1,
                              sheet.max_row - last_row_with_content)

        # Criar uma cópia das linhas para evitar alterar a estrutura durante a iteração
        rows_to_delete = []

        
        remover_espacos(int(dados_config[4]), int(dados_config[5]))

        # Remover as linhas indesejadas
        for row in rows_to_delete:
            sheet.delete_rows(row[0].row, 1)

    workbook.save(arquivo_saida)
    workbook.close()

    main.progressBar.hide()

    main.bt_process.setText('REMOVER ESPAÇOS EM BRANCO NA PLANILHA')

    alert('Espaços em branco removidos com Sucesso!  ')


# EltonTOC
def split():
    texto = main.input_text.toPlainText()
    resultado = ' '.join(texto.split())
    main.input_text.setText(resultado)
  

def reset_toc():
    main.input_text.setText('')


def copy():
    clipboard.copy(main.input_text.toPlainText())


# interface
app = QtWidgets.QApplication([])
# main
main = uic.loadUi('arquivos/ui/main.ui')
edit = uic.loadUi('arquivos/ui/edit.ui')
main.bt_process.clicked.connect(processar)
main.progressBar.hide()
main.actionOpen.triggered.connect(open_file)
main.bt_open_plan.clicked.connect(open_file)
main.bt_edit_plan.clicked.connect(save_file)
main.bt_config.clicked.connect(editar_configuracao)
main.bt_reset.clicked.connect(reset)
edit.bt_save_config.clicked.connect(salvar_configuracao)
main.actionSalvarArquivo.triggered.connect(save_file)
main.bt_split.clicked.connect(split)
main.bt_copy.clicked.connect(copy)
main.bt_reset_toc.clicked.connect(reset_toc)
# inicializar
main.show()
app.exec()
